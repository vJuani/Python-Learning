"""
Smoke tests for operation summary HTML / PDF / Excel.
"""

from __future__ import annotations

import io
import os
import tempfile
import unittest
from pathlib import Path

_TEST_TMP = tempfile.TemporaryDirectory()
_PRIVATE_ROOT = Path(_TEST_TMP.name) / "uploads"
_PRIVATE_ROOT.mkdir(parents=True, exist_ok=True)
os.environ["PRIVATE_UPLOAD_ROOT"] = str(_PRIVATE_ROOT)
os.environ["DATABASE_PATH"] = str(
    Path(_TEST_TMP.name) / "test_operation_summary.db"
)

from openpyxl import load_workbook

from modules.access_codes import hash_access_secret
from modules.auth import ROLE_ADMIN, ROLE_AGENT, hash_password
from modules.config import apply_config
from modules.database import add_agent, add_user, create_tables
from modules.database.connection import get_connection
from modules.database.guest_access_repository import (
    create_guest_access,
    get_guest_access_by_token_hash,
)
from modules.database.organization_settings_repository import (
    update_organization_settings,
)
from modules.database.organizations_repository import add_organization
from modules.excel_operation_summary import build_operation_summary_xlsx
from modules.operation_summary import load_operation_summary
from modules.pdf_operation_summary import build_operation_summary_pdf
from modules.vat_documents import (
    DOC_TYPE_MARTILLERO_CLIENT,
    upload_or_replace_vat_document,
)
from web_app import app


class FakeStorage:
    def __init__(self, filename, data, mimetype="application/pdf"):
        self.filename = filename
        self.stream = io.BytesIO(data)
        self.mimetype = mimetype

    def save(self, path):
        Path(path).write_bytes(self.stream.getvalue())
        self.stream.seek(0)


def _pdf_bytes():
    return b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n" + b"0" * 120


class OperationSummaryTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        apply_config(app)
        app.config["TESTING"] = True
        create_tables()

        cls.org = add_organization("Summary Org")
        cls.org_no_logo = add_organization("No Logo Org")

        # Brand logo file for org with logo.
        logo_dir = (
            Path(app.root_path)
            / "static"
            / "uploads"
            / "organizations"
            / str(cls.org)
        )
        logo_dir.mkdir(parents=True, exist_ok=True)
        cls.logo_rel = (
            f"uploads/organizations/{cls.org}/logo-test.png"
        )
        # Minimal PNG
        (Path(app.root_path) / "static" / cls.logo_rel).write_bytes(
            b"\x89PNG\r\n\x1a\n"
            + b"\x00" * 64
        )
        update_organization_settings(
            cls.org,
            display_name="Summary Org Display",
            default_language="es",
            default_currency="USD",
            timezone="America/Argentina/Buenos_Aires",
            logo_path=cls.logo_rel,
            accent_color=None,
        )

        cls.agent = add_agent("Summary Agent", "Alto", cls.org)
        cls.other_agent = add_agent(
            "Other Agent",
            "Puro",
            cls.org,
        )
        pwd = hash_password("Password1")
        cls.admin = add_user(
            "summary_admin",
            pwd,
            ROLE_ADMIN,
            cls.org,
            email="summary_admin@example.com",
        )
        cls.user_agent = add_user(
            "summary_agent",
            pwd,
            ROLE_AGENT,
            cls.org,
            agent_id=cls.agent,
            email="summary_agent@example.com",
        )
        cls.user_other = add_user(
            "summary_other",
            pwd,
            ROLE_AGENT,
            cls.org,
            agent_id=cls.other_agent,
            email="summary_other@example.com",
        )

        cls.op_usd = cls._insert_operation(
            cls.org,
            cls.agent,
            currency="USD",
            status="approved",
            created_by=cls.admin,
            reviewed_by=cls.admin,
        )
        cls.op_ars = cls._insert_operation(
            cls.org,
            cls.agent,
            currency="ARS",
            status="pending",
            created_by=cls.user_agent,
            reviewed_by=None,
            exchange_rate=1200,
            original_amount=120000000,
            sale_price=100000,
        )
        cls.op_rejected = cls._insert_operation(
            cls.org,
            cls.agent,
            currency="USD",
            status="rejected",
            created_by=cls.user_agent,
            reviewed_by=cls.admin,
            rejection_reason="Incomplete data",
        )
        cls.op_other = cls._insert_operation(
            cls.org,
            cls.other_agent,
            currency="USD",
            status="approved",
            created_by=cls.user_other,
            reviewed_by=cls.admin,
        )

        storage = FakeStorage("invoice.pdf", _pdf_bytes())
        upload_or_replace_vat_document(
            organization_id=cls.org,
            operation_id=cls.op_usd,
            doc_type=DOC_TYPE_MARTILLERO_CLIENT,
            file_storage=storage,
            uploaded_by_user_id=cls.admin,
        )

        cls.guest_hash = hash_access_secret("summary-guest")
        create_guest_access(
            cls.org,
            cls.guest_hash,
            cls.admin,
            label="summary-guest",
        )

    @classmethod
    def tearDownClass(cls):
        _TEST_TMP.cleanup()

    @classmethod
    def _insert_operation(
        cls,
        organization_id,
        agent_id,
        *,
        currency,
        status,
        created_by,
        reviewed_by,
        exchange_rate=1,
        original_amount=100000,
        sale_price=100000,
        rejection_reason=None,
    ):
        connection = get_connection()
        cursor = connection.cursor()
        cursor.execute(
            """
            INSERT INTO properties (
                address, jurisdiction, organization_id, agent_id, status
            )
            VALUES (?, 'CABA', ?, ?, 'approved')
            """,
            (
                f"Addr {organization_id}-{agent_id}-{status}",
                organization_id,
                agent_id,
            ),
        )
        property_id = cursor.lastrowid
        cursor.execute(
            """
            INSERT INTO operations (
                operation_date, agent_id, property_id, organization_id,
                was_invoiced, vat_amount, sale_price, commission_rate,
                total_commission, commission_after_abao, abao, martillero,
                agent_payment, office_payment, office_total, currency,
                original_amount, exchange_rate, status, rejection_reason,
                created_by_user_id, reviewed_by_user_id, reviewed_at
            )
            VALUES (
                '15/03/2026', ?, ?, ?, 'yes',
                210, ?, 3, 3000, 2850, 150, 114,
                1641.6, 1094.4, 1454.4, ?,
                ?, ?, ?, ?,
                ?, ?, ?
            )
            """,
            (
                agent_id,
                property_id,
                organization_id,
                sale_price,
                currency,
                original_amount,
                exchange_rate,
                status,
                rejection_reason,
                created_by,
                reviewed_by,
                "2026-03-16T12:00:00" if reviewed_by else None,
            ),
        )
        operation_id = cursor.lastrowid
        connection.commit()
        connection.close()
        return operation_id

    def _client_as(self, user_id=None, guest=False, language="es"):
        client = app.test_client()
        with client.session_transaction() as sess:
            sess.clear()
            sess["language"] = language
            if guest:
                access = get_guest_access_by_token_hash(
                    self.guest_hash
                )
                sess["guest_access_id"] = access["id"]
                sess["guest_organization_id"] = access[
                    "organization_id"
                ]
                sess["guest_token_hash"] = self.guest_hash
            elif user_id is not None:
                sess["user_id"] = user_id
        return client

    def test_admin_html_usd_with_docs_es(self):
        client = self._client_as(self.admin, language="es")
        response = client.get(
            f"/operations/{self.op_usd}/summary"
        )
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("Resumen de operación", html)
        self.assertIn("invoice.pdf", html)
        self.assertIn("Descargar PDF", html)
        self.assertIn("Descargar Excel", html)

    def test_admin_html_ars_pending_en(self):
        client = self._client_as(self.admin, language="en")
        response = client.get(
            f"/operations/{self.op_ars}/summary"
        )
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("Operation summary", html)
        self.assertIn("Pending", html)
        self.assertIn("ARS", html)
        self.assertIn("Download PDF", html)

    def test_rejected_operation(self):
        client = self._client_as(self.admin)
        response = client.get(
            f"/operations/{self.op_rejected}/summary"
        )
        html = response.get_data(as_text=True)
        self.assertIn("Incomplete data", html)
        self.assertIn("Rechazada", html)

    def test_agent_own_and_denied_other(self):
        client = self._client_as(self.user_agent)
        ok = client.get(f"/operations/{self.op_usd}/summary")
        self.assertEqual(ok.status_code, 200)

        denied = client.get(
            f"/operations/{self.op_other}/summary"
        )
        self.assertEqual(denied.status_code, 302)
        self.assertEqual(denied.headers.get("Location"), "/")

    def test_guest_limited_no_docs(self):
        client = self._client_as(guest=True)
        response = client.get(
            f"/operations/{self.op_usd}/summary"
        )
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertNotIn("invoice.pdf", html)
        self.assertIn(
            "Documentación privada no disponible",
            html,
        )

        pdf = client.get(
            f"/operations/{self.op_usd}/summary/pdf"
        )
        self.assertEqual(pdf.status_code, 200)
        self.assertTrue(
            pdf.data.startswith(b"%PDF")
        )
        self.assertNotIn(b"invoice.pdf", pdf.data)

        xlsx = client.get(
            f"/operations/{self.op_usd}/summary/xlsx"
        )
        self.assertEqual(xlsx.status_code, 200)
        workbook = load_workbook(io.BytesIO(xlsx.data))
        self.assertNotIn("Documentación", workbook.sheetnames)

    def test_pdf_download_admin_with_logo(self):
        client = self._client_as(self.admin, language="es")
        response = client.get(
            f"/operations/{self.op_usd}/summary/pdf"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/pdf",
        )
        self.assertTrue(response.data.startswith(b"%PDF"))
        disposition = response.headers.get(
            "Content-Disposition",
            "",
        )
        self.assertIn(
            "operation-summary_es.pdf",
            disposition,
        )

    def test_xlsx_numeric_cells_and_sheets(self):
        client = self._client_as(self.admin, language="es")
        response = client.get(
            f"/operations/{self.op_usd}/summary/xlsx"
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data))
        self.assertEqual(
            workbook.sheetnames,
            ["Resumen", "Comisión", "Aprobación", "Documentación"],
        )
        commission = workbook["Comisión"]
        # Banner + headers occupy rows 1-4; data starts at row 5.
        # Row 6 = total commission amount (numeric).
        total_cell = commission.cell(row=6, column=2)
        self.assertIsInstance(total_cell.value, (int, float))
        self.assertNotIsInstance(total_cell.value, str)

        docs = workbook["Documentación"]
        self.assertEqual(docs.cell(row=5, column=2).value, "invoice.pdf")

    def test_operation_without_documents_sheet(self):
        client = self._client_as(self.admin)
        response = client.get(
            f"/operations/{self.op_ars}/summary/xlsx"
        )
        workbook = load_workbook(io.BytesIO(response.data))
        docs = workbook["Documentación"]
        self.assertIn(
            "Sin documentos",
            str(docs.cell(row=5, column=1).value),
        )

    def test_summary_without_org_logo_fallback(self):
        from modules.database import get_operation_record

        # Build summary for org with logo removed via settings override.
        operation = get_operation_record(self.op_usd, self.org)
        summary = load_operation_summary(
            operation,
            language="en",
            can_see_documents=True,
        )
        # Force brand without org logo path readability
        summary["brand"]["logo_path"] = None
        pdf_bytes = build_operation_summary_pdf(summary)
        self.assertTrue(pdf_bytes.startswith(b"%PDF"))

        xlsx_bytes = build_operation_summary_xlsx(summary)
        self.assertGreater(len(xlsx_bytes), 100)

    def test_no_internal_ids_in_exports(self):
        client = self._client_as(self.admin)
        pdf = client.get(
            f"/operations/{self.op_usd}/summary/pdf"
        ).data
        xlsx = client.get(
            f"/operations/{self.op_usd}/summary/xlsx"
        ).data
        for blob in (pdf, xlsx):
            self.assertNotIn(b"organization_id", blob)
            self.assertNotIn(b"agent_db_id", blob)
            self.assertNotIn(b"stored_name", blob)
            self.assertNotIn(str(_PRIVATE_ROOT).encode(), blob)


if __name__ == "__main__":
    unittest.main()

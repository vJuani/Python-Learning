"""Tests for organization-wide reports presets and exports."""

from __future__ import annotations

import calendar
import io
import os
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

_TEST_TMP = tempfile.TemporaryDirectory()
os.environ["DATABASE_PATH"] = str(
    Path(_TEST_TMP.name) / "test_organization_reports.db"
)
os.environ["PRIVATE_UPLOAD_ROOT"] = str(
    Path(_TEST_TMP.name) / "uploads"
)

from openpyxl import load_workbook

from modules.auth import ROLE_ADMIN, ROLE_AGENT, hash_password
from modules.config import apply_config
from modules.database import add_agent, add_user, create_tables
from modules.database.connection import get_connection
from modules.database.organizations_repository import add_organization
from modules.excel_organization_report import (
    build_organization_report_xlsx,
)
from modules.organization_reports import (
    build_download_basename,
    load_organization_report,
    parse_report_filters,
)
from modules.pdf_organization_report import (
    build_organization_report_pdf,
)
from web_app import app


class OrganizationReportsTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        apply_config(app)
        app.config["TESTING"] = True
        create_tables()

        cls.org = add_organization("Reports Org")
        cls.other_org = add_organization("Other Reports Org")
        cls.agent_a = add_agent("Agent A", "Alto", cls.org)
        cls.agent_b = add_agent("Agent B", "Puro", cls.org)
        cls.other_org_agent = add_agent(
            "Other Org Agent",
            "Junior",
            cls.other_org,
        )

        pwd = hash_password("Password1")
        cls.admin = add_user(
            "reports_admin",
            pwd,
            ROLE_ADMIN,
            cls.org,
            email="reports_admin@example.com",
        )
        cls.agent_user = add_user(
            "reports_agent",
            pwd,
            ROLE_AGENT,
            cls.org,
            agent_id=cls.agent_a,
            email="reports_agent@example.com",
        )

        now = datetime.utcnow()
        cls.year = now.year
        cls.month = now.month
        last_day = calendar.monthrange(cls.year, cls.month)[1]
        cls.date_in_month = (
            f"15/{cls.month:02d}/{cls.year:04d}"
        )
        cls.date_prev = (
            f"10/{cls.month:02d}/{cls.year - 1:04d}"
            if cls.month != 1
            else f"10/12/{cls.year - 1:04d}"
        )

        cls.prop_a = cls._insert_property(
            cls.org, cls.agent_a, "Street A"
        )
        cls.prop_b = cls._insert_property(
            cls.org, cls.agent_b, "Street B"
        )
        cls.prop_other = cls._insert_property(
            cls.other_org,
            cls.other_org_agent,
            "Other Street",
        )

        cls.op_approved_invoiced = cls._insert_operation(
            cls.org,
            cls.agent_a,
            cls.prop_a,
            operation_date=cls.date_in_month,
            was_invoiced="yes",
            status="approved",
            sale_price=100000,
            total_commission=3000,
            agent_payment=1800,
            office_payment=1200,
            vat_amount=210,
        )
        cls.op_approved_not_invoiced = cls._insert_operation(
            cls.org,
            cls.agent_a,
            cls.prop_a,
            operation_date=cls.date_in_month,
            was_invoiced="no",
            status="approved",
            sale_price=50000,
            total_commission=1500,
            agent_payment=900,
            office_payment=600,
            vat_amount=0,
        )
        cls.op_pending = cls._insert_operation(
            cls.org,
            cls.agent_b,
            cls.prop_b,
            operation_date=cls.date_in_month,
            was_invoiced="yes",
            status="pending",
            sale_price=200000,
            total_commission=6000,
            agent_payment=3600,
            office_payment=2400,
            vat_amount=0,
        )
        cls.op_other_month = cls._insert_operation(
            cls.org,
            cls.agent_b,
            cls.prop_b,
            operation_date=cls.date_prev,
            was_invoiced="no",
            status="approved",
            sale_price=80000,
            total_commission=2400,
            agent_payment=1440,
            office_payment=960,
            vat_amount=0,
        )
        cls.op_other_org = cls._insert_operation(
            cls.other_org,
            cls.other_org_agent,
            cls.prop_other,
            operation_date=cls.date_in_month,
            was_invoiced="yes",
            status="approved",
            sale_price=999999,
            total_commission=99999,
            agent_payment=1,
            office_payment=1,
            vat_amount=1,
        )

        cls.client = app.test_client()

    @classmethod
    def _insert_property(cls, org_id, agent_id, address):
        connection = get_connection()
        cursor = connection.cursor()
        cursor.execute(
            """
            INSERT INTO properties (
                address,
                jurisdiction,
                organization_id,
                agent_id,
                status
            )
            VALUES (?, 'CABA', ?, ?, 'approved')
            """,
            (address, org_id, agent_id),
        )
        property_id = cursor.lastrowid
        connection.commit()
        connection.close()
        return property_id

    @classmethod
    def _insert_operation(
        cls,
        org_id,
        agent_id,
        property_id,
        *,
        operation_date,
        was_invoiced,
        status,
        sale_price,
        total_commission,
        agent_payment,
        office_payment,
        vat_amount,
    ):
        connection = get_connection()
        cursor = connection.cursor()
        cursor.execute(
            """
            INSERT INTO operations (
                operation_date,
                agent_id,
                property_id,
                organization_id,
                was_invoiced,
                vat_amount,
                sale_price,
                commission_rate,
                total_commission,
                commission_after_abao,
                abao,
                martillero,
                agent_payment,
                office_payment,
                office_total,
                currency,
                original_amount,
                exchange_rate,
                status,
                created_by_user_id
            )
            VALUES (
                ?, ?, ?, ?, ?, ?, ?, 3, ?, ?, 0, 0,
                ?, ?, ?, 'USD', ?, 1, ?, ?
            )
            """,
            (
                operation_date,
                agent_id,
                property_id,
                org_id,
                was_invoiced,
                vat_amount,
                sale_price,
                total_commission,
                total_commission,
                agent_payment,
                office_payment,
                office_payment + vat_amount,
                sale_price,
                status,
                cls.admin,
            ),
        )
        operation_id = cursor.lastrowid
        connection.commit()
        connection.close()
        return operation_id

    def _login(self, username):
        return self.client.post(
            "/login",
            data={
                "username": username,
                "password": "Password1",
            },
            follow_redirects=True,
        )

    def test_parse_year_mode_shows_full_year_bounds(self):
        errors, parsed = parse_report_filters(
            {
                "view": "general",
                "period_mode": "year",
                "year": "2026",
            }
        )
        self.assertEqual(errors, [])
        self.assertEqual(parsed["period_mode"], "year")
        self.assertEqual(parsed["query_filters"]["date_from"], "20260101")
        self.assertEqual(parsed["query_filters"]["date_to"], "20261231")
        self.assertEqual(parsed["form"]["month"], "")

    def test_month_options_use_names(self):
        from modules.organization_reports import month_options

        es = month_options("es")
        en = month_options("en")
        self.assertEqual(es[0]["label"], "Enero")
        self.assertEqual(en[7]["label"], "August")

    def test_parse_uninvoiced_forces_approved_and_not_invoiced(self):
        errors, parsed = parse_report_filters(
            {"view": "uninvoiced"}
        )
        self.assertEqual(errors, [])
        self.assertEqual(
            parsed["query_filters"]["was_invoiced"],
            "no",
        )
        self.assertEqual(
            parsed["query_filters"]["status"],
            "approved",
        )

    def test_parse_monthly_sets_month_bounds(self):
        errors, parsed = parse_report_filters(
            {
                "view": "monthly",
                "year": str(self.year),
                "month": str(self.month),
            }
        )
        self.assertEqual(errors, [])
        last_day = calendar.monthrange(self.year, self.month)[1]
        self.assertEqual(
            parsed["query_filters"]["date_from"],
            f"{self.year:04d}{self.month:02d}01",
        )
        self.assertEqual(
            parsed["query_filters"]["date_to"],
            f"{self.year:04d}{self.month:02d}{last_day:02d}",
        )

    def test_general_defaults_to_approved_metrics(self):
        report = load_organization_report(
            self.org,
            {"view": "general"},
            language="es",
        )
        self.assertEqual(report["errors"], [])
        # pending op excluded from monetary metrics
        self.assertEqual(report["metrics"]["operations_count"], 3)
        self.assertEqual(
            report["metrics"]["total_commission"],
            3000 + 1500 + 2400,
        )
        self.assertEqual(report["metrics"]["approved_count"], 3)
        self.assertEqual(report["metrics"]["pending_count"], 1)
        self.assertTrue(
            all(
                item["id"].startswith("COM-")
                for item in report["operations"]
            )
        )
        # other org excluded
        self.assertNotIn(
            99999,
            [
                item["total_commission"]
                for item in report["operations"]
            ],
        )

    def test_monthly_preset_filters_current_month(self):
        report = load_organization_report(
            self.org,
            {
                "view": "monthly",
                "year": str(self.year),
                "month": str(self.month),
            },
            language="es",
        )
        self.assertEqual(report["metrics"]["operations_count"], 2)
        self.assertEqual(
            report["metrics"]["invoiced_count"],
            1,
        )
        self.assertEqual(
            report["metrics"]["not_invoiced_count"],
            1,
        )
        self.assertEqual(
            report["download_basename"],
            f"reporte_mensual_{self.year}-{self.month:02d}",
        )

    def test_invoiced_preset(self):
        report = load_organization_report(
            self.org,
            {"view": "invoiced"},
            language="es",
        )
        self.assertEqual(report["metrics"]["operations_count"], 1)
        self.assertEqual(
            report["operations"][0]["was_invoiced"],
            "yes",
        )

    def test_uninvoiced_preset(self):
        report = load_organization_report(
            self.org,
            {"view": "uninvoiced"},
            language="es",
        )
        self.assertEqual(report["metrics"]["operations_count"], 2)
        self.assertTrue(
            all(
                item["was_invoiced"] == "no"
                and item["status"] == "approved"
                for item in report["operations"]
            )
        )

    def test_agent_scope_limits_operations(self):
        report = load_organization_report(
            self.org,
            {"view": "general"},
            language="es",
            scoped_agent_id=self.agent_a,
        )
        self.assertEqual(report["metrics"]["operations_count"], 2)
        self.assertTrue(
            all(
                item["agent"] == "Agent A"
                for item in report["operations"]
            )
        )

    def test_pdf_and_excel_exports(self):
        report = load_organization_report(
            self.org,
            {
                "view": "monthly",
                "year": str(self.year),
                "month": str(self.month),
            },
            language="es",
        )
        pdf_bytes = build_organization_report_pdf(report)
        self.assertTrue(pdf_bytes.startswith(b"%PDF"))

        xlsx_bytes = build_organization_report_xlsx(report)
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        self.assertEqual(
            workbook.sheetnames,
            ["Summary", "By month", "Agents ranking", "Operations"],
        )
        summary = workbook["Summary"]
        # operations count cell should be numeric
        self.assertEqual(summary["B7"].value, 2)
        self.assertIsInstance(summary["B10"].value, (int, float))
        self.assertEqual(summary["B10"].value, 150000)

    def test_download_basename_en_monthly(self):
        errors, parsed = parse_report_filters(
            {
                "view": "monthly",
                "year": "2026",
                "month": "8",
            }
        )
        self.assertEqual(errors, [])
        self.assertEqual(
            build_download_basename(parsed, "en"),
            "monthly_report_2026-08",
        )

    def test_reports_routes_admin(self):
        self._login("reports_admin")
        response = self.client.get("/reports?view=general")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Reportes", response.data)

        pdf = self.client.get("/reports/pdf?view=uninvoiced")
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf.mimetype, "application/pdf")

        xlsx = self.client.get("/reports/xlsx?view=invoiced")
        self.assertEqual(xlsx.status_code, 200)
        self.assertIn(
            "spreadsheetml",
            xlsx.mimetype,
        )


if __name__ == "__main__":
    unittest.main()

"""
Excel export for organization reports (openpyxl).
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


NAVY = "001838"
ACCENT = "0860C8"
ACCENT_SOFT = "E8F1FC"
ROW_ALT = "F5F8FC"
LABEL_BG = "EEF3F9"
MONEY_BG = "EAF6FF"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=10, color="D6E4F5")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=NAVY)
VALUE_FONT = Font(name="Calibri", size=11, color=NAVY)
MONEY_FONT = Font(name="Calibri", size=11, color=ACCENT)

THIN = Side(style="thin", color="C9D6E5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FORMAT = "#,##0.00"
RATE_FORMAT = "#,##0.0000"


def _fill(hex_color):
    return PatternFill(
        start_color=hex_color,
        end_color=hex_color,
        fill_type="solid",
    )


def _to_number(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _write_banner(worksheet, title, subtitle, merge_columns):
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=merge_columns,
    )
    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=merge_columns,
    )
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.fill = _fill(NAVY)
    title_cell.alignment = Alignment(vertical="center")
    subtitle_cell = worksheet.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.fill = _fill(ACCENT)
    worksheet.row_dimensions[1].height = 28
    worksheet.row_dimensions[2].height = 18


def _autosize(worksheet, min_width=12, max_width=36):
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            length = max(length, len(str(cell.value)))

        worksheet.column_dimensions[letter].width = min(
            max(length + 2, min_width),
            max_width,
        )


def _header_row(worksheet, row, values):
    for index, value in enumerate(values, start=1):
        cell = worksheet.cell(row=row, column=index, value=value)
        cell.font = HEADER_FONT
        cell.fill = _fill(NAVY)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")


def _money_cell(worksheet, row, column, value):
    cell = worksheet.cell(
        row=row,
        column=column,
        value=_to_number(value),
    )
    cell.number_format = MONEY_FORMAT
    cell.font = MONEY_FONT
    cell.fill = _fill(MONEY_BG)
    cell.border = BORDER
    return cell


def _text_cell(worksheet, row, column, value, alt=False):
    cell = worksheet.cell(row=row, column=column, value=value)
    cell.font = VALUE_FONT
    cell.border = BORDER

    if alt:
        cell.fill = _fill(ROW_ALT)

    return cell


def _write_summary_sheet(workbook, report):
    labels = report["labels"]
    metrics = report["metrics"]
    worksheet = workbook.create_sheet(labels["sheet_summary"][:31], 0)
    org = report["brand"].get("organization_name") or ""
    subtitle = report["brand"]["app_name"] + (f" — {org}" if org else "")
    _write_banner(worksheet, labels["title"], subtitle, 2)

    worksheet.cell(row=4, column=1, value=labels["filters_applied"]).font = (
        LABEL_FONT
    )
    filters = " · ".join(report["active_filters"]) or labels["empty"]
    worksheet.cell(row=4, column=2, value=filters).font = VALUE_FONT
    worksheet.cell(row=5, column=1, value=labels["official_note"]).font = (
        VALUE_FONT
    )

    rows = [
        (labels["operations_count"], metrics["operations_count"], False),
        (labels["approved_count"], metrics["approved_count"], False),
        (labels["pending_count"], metrics["pending_count"], False),
        (labels["volume_usd"], metrics["volume_usd"], True),
        (labels["total_commission"], metrics["total_commission"], True),
        (labels["agent_payments"], metrics["agent_payments"], True),
        (labels["office_net"], metrics["office_net"], True),
        (labels["vat_total"], metrics["vat_total"], True),
        (labels["properties_count"], metrics["properties_count"], False),
        (labels["average_commission"], metrics["average_commission"], True),
        (labels["invoiced_count"], metrics["invoiced_count"], False),
        (labels["not_invoiced_count"], metrics["not_invoiced_count"], False),
    ]

    start = 7

    for offset, (label, value, is_money) in enumerate(rows):
        row = start + offset
        label_cell = worksheet.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.fill = _fill(LABEL_BG)
        label_cell.border = BORDER

        if is_money:
            _money_cell(worksheet, row, 2, value)
        else:
            cell = worksheet.cell(
                row=row,
                column=2,
                value=int(value) if isinstance(value, float) and value == int(value) else value,
            )
            cell.font = VALUE_FONT
            cell.border = BORDER
            cell.fill = _fill(ACCENT_SOFT)

    _autosize(worksheet)
    return worksheet


def _write_monthly_sheet(workbook, report):
    labels = report["labels"]
    worksheet = workbook.create_sheet(labels["sheet_by_month"][:31])
    _write_banner(
        worksheet,
        labels["monthly_title"],
        report["brand"]["app_name"],
        6,
    )
    _header_row(
        worksheet,
        4,
        [
            labels["col_month"],
            labels["col_operations"],
            labels["col_volume"],
            labels["col_commission"],
            labels["col_agent_payment"],
            labels["col_office_net"],
        ],
    )

    for index, item in enumerate(report["monthly_series"], start=5):
        alt = index % 2 == 0
        _text_cell(worksheet, index, 1, item["month"], alt)
        cell = _text_cell(
            worksheet,
            index,
            2,
            item["operations_count"],
            alt,
        )
        cell.alignment = Alignment(horizontal="right")
        _money_cell(worksheet, index, 3, item["volume_usd"])
        _money_cell(worksheet, index, 4, item["total_commission"])
        _money_cell(worksheet, index, 5, item["agent_payments"])
        _money_cell(worksheet, index, 6, item["office_net"])

    if not report["monthly_series"]:
        _text_cell(worksheet, 5, 1, labels["no_operations"])

    _autosize(worksheet)


def _write_agents_sheet(workbook, report):
    labels = report["labels"]
    worksheet = workbook.create_sheet(labels["sheet_agents"][:31])
    _write_banner(
        worksheet,
        labels["ranking_title"],
        report["brand"]["app_name"],
        7,
    )
    _header_row(
        worksheet,
        4,
        [
            labels["col_rank"],
            labels["col_agent"],
            labels["col_agent_type"],
            labels["col_operations"],
            labels["col_commission"],
            labels["col_agent_payment"],
            labels["col_volume"],
        ],
    )

    for index, item in enumerate(report["agent_ranking"], start=5):
        alt = index % 2 == 0
        _text_cell(worksheet, index, 1, item["rank"], alt)
        _text_cell(worksheet, index, 2, item["agent_name"], alt)
        _text_cell(worksheet, index, 3, item["agent_type"], alt)
        _text_cell(worksheet, index, 4, item["operations_count"], alt)
        _money_cell(worksheet, index, 5, item["total_commission"])
        _money_cell(worksheet, index, 6, item["agent_payments"])
        _money_cell(worksheet, index, 7, item["volume_usd"])

    if not report["agent_ranking"]:
        _text_cell(worksheet, 5, 1, labels["no_operations"])

    _autosize(worksheet)


def _write_operations_sheet(workbook, report):
    labels = report["labels"]
    worksheet = workbook.create_sheet(labels["sheet_operations"][:31])
    _write_banner(
        worksheet,
        labels["operations_title"],
        report["brand"]["app_name"],
        16,
    )
    _header_row(
        worksheet,
        4,
        [
            labels["col_id"],
            labels["col_date"],
            labels["col_agent"],
            labels["col_agent_type"],
            labels["col_property"],
            labels["col_jurisdiction"],
            labels["col_status"],
            labels["col_invoiced"],
            labels["col_currency"],
            labels["col_original_amount"],
            labels["col_exchange_rate"],
            labels["col_sale_price_usd"],
            labels["col_commission"],
            labels["col_agent_payment"],
            labels["col_office_net"],
            labels["col_vat"],
        ],
    )

    for index, item in enumerate(report["operations"], start=5):
        alt = index % 2 == 0
        _text_cell(worksheet, index, 1, item["id"], alt)
        _text_cell(worksheet, index, 2, item["date"], alt)
        _text_cell(worksheet, index, 3, item["agent"], alt)
        _text_cell(worksheet, index, 4, item["agent_type"], alt)
        _text_cell(worksheet, index, 5, item["property"], alt)
        _text_cell(worksheet, index, 6, item["jurisdiction"], alt)
        _text_cell(worksheet, index, 7, item["status_label"], alt)
        _text_cell(worksheet, index, 8, item["was_invoiced_label"], alt)
        _text_cell(worksheet, index, 9, item["currency"], alt)

        original = worksheet.cell(
            row=index,
            column=10,
            value=_to_number(item["original_amount"]),
        )
        original.number_format = MONEY_FORMAT
        original.font = VALUE_FONT
        original.border = BORDER

        if alt:
            original.fill = _fill(ROW_ALT)

        rate = worksheet.cell(
            row=index,
            column=11,
            value=_to_number(item["exchange_rate"]),
        )
        rate.number_format = RATE_FORMAT
        rate.font = VALUE_FONT
        rate.border = BORDER

        if alt:
            rate.fill = _fill(ROW_ALT)

        _money_cell(worksheet, index, 12, item["sale_price"])
        _money_cell(worksheet, index, 13, item["total_commission"])
        _money_cell(worksheet, index, 14, item["agent_payment"])
        _money_cell(worksheet, index, 15, item["office_payment"])
        _money_cell(worksheet, index, 16, item["vat_amount"])

    if not report["operations"]:
        _text_cell(worksheet, 5, 1, labels["no_operations"])

    _autosize(worksheet, min_width=10, max_width=28)


def build_organization_report_xlsx(report):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _write_summary_sheet(workbook, report)
    _write_monthly_sheet(workbook, report)
    _write_agents_sheet(workbook, report)
    _write_operations_sheet(workbook, report)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

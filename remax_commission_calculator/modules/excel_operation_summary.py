"""
Excel export for operation summaries (openpyxl).
"""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# Brand palette (Commission Calculator)
NAVY = "001838"
ACCENT = "0860C8"
ACCENT_SOFT = "E8F1FC"
ROW_ALT = "F5F8FC"
ROW_BASE = "FFFFFF"
LABEL_BG = "EEF3F9"
MONEY_BG = "EAF6FF"
TOTAL_BG = "D7EBFF"
STATUS_APPROVED = "DCF5E5"
STATUS_PENDING = "FFF4D6"
STATUS_REJECTED = "FDE2E1"
STATUS_DRAFT = "E8EEF5"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=10, color="D6E4F5")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=NAVY)
VALUE_FONT = Font(name="Calibri", size=11, color=NAVY)
MONEY_FONT = Font(name="Calibri", size=11, color=ACCENT)
TOTAL_FONT = Font(name="Calibri", size=12, bold=True, color=NAVY)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=WHITE)

THIN = Side(style="thin", color="C9D6E5")
BORDER = Border(
    left=THIN,
    right=THIN,
    top=THIN,
    bottom=THIN,
)

MONEY_FORMAT = "#,##0.00"
RATE_FORMAT = "#,##0.0000"
PERCENT_FORMAT = "0.00"


def _fill(hex_color):
    return PatternFill(
        start_color=hex_color,
        end_color=hex_color,
        fill_type="solid",
    )


def _to_excel_number(value):
    if value is None:
        return None

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _status_fill(status):
    mapping = {
        "approved": STATUS_APPROVED,
        "pending": STATUS_PENDING,
        "rejected": STATUS_REJECTED,
        "draft": STATUS_DRAFT,
    }
    return _fill(mapping.get(status, ROW_ALT))


def _write_banner(worksheet, title, subtitle, merge_columns=2):
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=merge_columns,
    )
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.fill = _fill(NAVY)
    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    worksheet.row_dimensions[1].height = 28

    for column in range(1, merge_columns + 1):
        cell = worksheet.cell(row=1, column=column)
        cell.fill = _fill(NAVY)
        cell.border = BORDER

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=merge_columns,
    )
    subtitle_cell = worksheet.cell(
        row=2,
        column=1,
        value=subtitle,
    )
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.fill = _fill(ACCENT)
    subtitle_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    worksheet.row_dimensions[2].height = 20

    for column in range(1, merge_columns + 1):
        cell = worksheet.cell(row=2, column=column)
        cell.fill = _fill(ACCENT)
        cell.border = BORDER


def _write_column_headers(worksheet, headers, row=4):
    for index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=index, value=header)
        cell.font = HEADER_FONT
        cell.fill = _fill(ACCENT)
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )
        cell.border = BORDER

    worksheet.row_dimensions[row].height = 22


def _style_value_cell(cell, kind, emphasize=False):
    cell.font = VALUE_FONT
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    cell.border = BORDER

    if kind == "money":
        cell.font = MONEY_FONT if not emphasize else TOTAL_FONT
        cell.fill = _fill(TOTAL_BG if emphasize else MONEY_BG)
        if cell.value is not None:
            cell.number_format = MONEY_FORMAT
    elif kind == "rate":
        cell.fill = _fill(ACCENT_SOFT)
        if cell.value is not None:
            cell.number_format = RATE_FORMAT
    elif kind == "percent":
        cell.fill = _fill(ACCENT_SOFT)
        if cell.value is not None:
            cell.number_format = PERCENT_FORMAT
    elif emphasize:
        cell.font = TOTAL_FONT
        cell.fill = _fill(TOTAL_BG)


def _write_pairs(
    worksheet,
    pairs,
    *,
    start_row,
    status=None,
    emphasize_labels=None,
):
    emphasize_labels = emphasize_labels or set()
    row = start_row

    for index, (label, value, kind) in enumerate(pairs):
        alt = index % 2 == 1
        label_cell = worksheet.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.fill = _fill(LABEL_BG if not alt else ROW_ALT)
        label_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )
        label_cell.border = BORDER

        value_cell = worksheet.cell(row=row, column=2)
        emphasize = label in emphasize_labels

        if kind in ("money", "rate", "percent"):
            number = _to_excel_number(value)
            value_cell.value = number
        else:
            value_cell.value = (
                value if value not in (None, "") else None
            )

        base_fill = ROW_ALT if alt else ROW_BASE
        value_cell.fill = _fill(base_fill)
        _style_value_cell(value_cell, kind, emphasize=emphasize)

        if status is not None and kind == "text" and index == 0:
            # First text row often is status on approval sheet;
            # only tint when label looks like status via emphasize set.
            pass

        if (
            status is not None
            and label.lower() in ("estado", "status")
        ):
            value_cell.fill = _status_fill(status)
            value_cell.font = Font(
                name="Calibri",
                size=11,
                bold=True,
                color=NAVY,
            )

        worksheet.row_dimensions[row].height = 20
        row += 1

    return row


def _autosize(worksheet, min_width=16, max_width=42):
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        worksheet.column_dimensions[
            get_column_letter(column)
        ].width = min(max(max_length + 3, min_width), max_width)


def _prepare_sheet(worksheet, title, subtitle, headers):
    _write_banner(worksheet, title, subtitle, merge_columns=2)
    _write_column_headers(worksheet, headers, row=4)
    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False


def build_operation_summary_xlsx(summary):
    labels = summary["labels"]
    operation = summary["operation"]
    numbers = summary["numbers"]
    people = summary["people"]
    brand = summary["brand"]
    empty = labels["empty_value"]

    org_name = brand.get("organization_name") or brand["app_name"]
    banner_sub = (
        f"{org_name}  ·  {operation['id']}  ·  "
        f"{operation['status_label']}"
    )

    workbook = Workbook()

    # --- Summary ---
    sheet_summary = workbook.active
    sheet_summary.title = labels["sheet_summary"][:31]
    _prepare_sheet(
        sheet_summary,
        f"{brand['app_name']} — {labels['sheet_summary']}",
        banner_sub,
        [
            "Campo" if summary["language"] == "es" else "Field",
            "Valor" if summary["language"] == "es" else "Value",
        ],
    )

    summary_pairs = [
        (labels["field_operation_id"], operation["id"], "text"),
        (labels["field_date"], operation["date"] or empty, "text"),
        (
            labels["field_status"],
            operation["status_label"],
            "text",
        ),
        (
            labels["field_agent"],
            operation["agent_name"] or empty,
            "text",
        ),
        (
            labels["field_property"],
            operation["property_address"] or empty,
            "text",
        ),
        (
            labels["field_property_id"],
            operation["property_id"] or empty,
            "text",
        ),
        (
            labels["field_jurisdiction"],
            operation["jurisdiction"] or empty,
            "text",
        ),
    ]

    for line in summary.get("billing_lines", []):
        summary_pairs.append(
            (line["label"], line["value"], line["kind"])
        )

    emphasize = {
        line["label"]
        for line in summary.get("billing_lines", [])
        if line.get("emphasize")
    }
    _write_pairs(
        sheet_summary,
        summary_pairs,
        start_row=5,
        status=operation["status"],
        emphasize_labels=emphasize,
    )
    _autosize(sheet_summary)

    # --- Commission ---
    sheet_commission = workbook.create_sheet(
        labels["sheet_commission"][:31]
    )
    _prepare_sheet(
        sheet_commission,
        f"{brand['app_name']} — {labels['sheet_commission']}",
        banner_sub,
        [
            "Campo" if summary["language"] == "es" else "Field",
            "Valor" if summary["language"] == "es" else "Value",
        ],
    )
    commission_pairs = [
        (line["label"], line["value"], line["kind"])
        for line in summary.get("commission_lines", [])
    ]
    commission_emphasize = {
        line["label"]
        for line in summary.get("commission_lines", [])
        if line.get("emphasize")
    }
    _write_pairs(
        sheet_commission,
        commission_pairs,
        start_row=5,
        emphasize_labels=commission_emphasize,
    )
    _autosize(sheet_commission)

    # --- Approval ---
    sheet_approval = workbook.create_sheet(
        labels["sheet_approval"][:31]
    )
    _prepare_sheet(
        sheet_approval,
        f"{brand['app_name']} — {labels['sheet_approval']}",
        banner_sub,
        [
            "Campo" if summary["language"] == "es" else "Field",
            "Valor" if summary["language"] == "es" else "Value",
        ],
    )
    approval_pairs = [
        (
            labels["field_status"],
            operation["status_label"],
            "text",
        ),
        (
            labels["field_created_by"],
            people["created_by_name"] or empty,
            "text",
        ),
        (
            labels["field_reviewed_by"],
            people["reviewed_by_name"] or empty,
            "text",
        ),
        (
            labels["field_reviewed_at"],
            operation["reviewed_at"] or empty,
            "text",
        ),
        (
            labels["field_rejection_reason"],
            operation["rejection_reason"] or empty,
            "text",
        ),
    ]
    _write_pairs(
        sheet_approval,
        approval_pairs,
        start_row=5,
        status=operation["status"],
    )
    _autosize(sheet_approval)

    # --- Documentation ---
    if summary["can_see_documents"]:
        sheet_docs = workbook.create_sheet(
            labels["sheet_documents"][:31]
        )
        _write_banner(
            sheet_docs,
            f"{brand['app_name']} — {labels['sheet_documents']}",
            banner_sub,
            merge_columns=4,
        )
        headers = [
            labels["field_doc_type"],
            labels["field_filename"],
            labels["field_uploaded_at"],
            labels.get(
                "field_uploaded_by",
                "Uploaded by",
            ),
        ]
        _write_column_headers(sheet_docs, headers, row=4)
        sheet_docs.freeze_panes = "A5"
        sheet_docs.sheet_view.showGridLines = False

        if summary["documents"]:
            for row_index, document in enumerate(
                summary["documents"],
                start=5,
            ):
                alt = (row_index % 2) == 0
                values = [
                    document["doc_type_label"],
                    document["original_filename"],
                    document["uploaded_at"],
                    document.get("uploaded_by_name") or empty,
                ]
                for column, value in enumerate(values, start=1):
                    cell = sheet_docs.cell(
                        row=row_index,
                        column=column,
                        value=value,
                    )
                    cell.font = VALUE_FONT
                    cell.fill = _fill(ROW_ALT if alt else ROW_BASE)
                    cell.border = BORDER
                    cell.alignment = Alignment(
                        vertical="center"
                    )
                sheet_docs.row_dimensions[row_index].height = 20
        else:
            cell = sheet_docs.cell(
                row=5,
                column=1,
                value=labels["no_documents"],
            )
            cell.font = VALUE_FONT
            cell.fill = _fill(ACCENT_SOFT)
            cell.border = BORDER
            sheet_docs.merge_cells("A5:D5")

        _autosize(sheet_docs, min_width=18, max_width=46)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


__all__ = [
    "build_operation_summary_xlsx",
]

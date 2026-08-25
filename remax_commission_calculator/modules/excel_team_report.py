"""
Excel export for team reports (openpyxl).
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


NAVY = "001838"
ACCENT = "0860C8"
ROW_ALT = "F5F8FC"
WHITE = "FFFFFF"
MONEY_FORMAT = "#,##0.00"

THIN = Side(style="thin", color="C9D6E5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color):
    return PatternFill(
        start_color=hex_color,
        end_color=hex_color,
        fill_type="solid",
    )


def _header_cell(cell, value):
    cell.value = value
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = _fill(NAVY)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _money_cell(cell, value):
    cell.value = float(value or 0)
    cell.number_format = MONEY_FORMAT
    cell.font = Font(name="Calibri", size=11, color=ACCENT)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="right")


def _text_cell(cell, value, alt=False):
    cell.value = value
    cell.font = Font(name="Calibri", size=11, color=NAVY)
    cell.border = BORDER
    if alt:
        cell.fill = _fill(ROW_ALT)


def build_team_report_xlsx(report):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    leader = report["leader"]
    metrics = report["metrics"]
    labels = report["labels"]

    summary["A1"] = labels["title"]
    summary["A1"].font = Font(
        name="Calibri", size=16, bold=True, color=WHITE
    )
    summary["A1"].fill = _fill(NAVY)
    summary.merge_cells("A1:B1")

    summary["A2"] = f"{leader['name']} ({leader['type']})"
    summary["A2"].font = Font(name="Calibri", size=11, color=NAVY)

    rows = [
        (labels["team_production"], metrics["team_production"], True),
        (labels["juniors_production"], metrics["juniors_production"], True),
        (labels["leader_own"], metrics["leader_own_income"], True),
        (labels["juniors_income"], metrics["juniors_income_to_leader"], True),
        (labels["combined"], metrics["combined_income"], True),
        (labels["operations"], metrics["operations_count"], False),
    ]

    start = 4
    for offset, (label, value, is_money) in enumerate(rows):
        row = start + offset
        _text_cell(summary.cell(row=row, column=1), label)
        cell = summary.cell(row=row, column=2)
        if is_money:
            _money_cell(cell, value)
        else:
            _text_cell(cell, value)

    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 18

    juniors = workbook.create_sheet("Juniors")
    headers = [
        "Junior",
        "Type",
        "Ops",
        "Production",
        "Junior yield",
        "TL income",
    ]
    for index, header in enumerate(headers, start=1):
        _header_cell(juniors.cell(row=1, column=index), header)

    for row_index, item in enumerate(report["junior_rows"], start=2):
        alt = row_index % 2 == 0
        _text_cell(
            juniors.cell(row=row_index, column=1),
            item["agent"]["name"],
            alt,
        )
        _text_cell(
            juniors.cell(row=row_index, column=2),
            item["agent"]["type"],
            alt,
        )
        _text_cell(
            juniors.cell(row=row_index, column=3),
            item["operations_count"],
            alt,
        )
        _money_cell(
            juniors.cell(row=row_index, column=4),
            item["production"],
        )
        _money_cell(
            juniors.cell(row=row_index, column=5),
            item["junior_yield"],
        )
        _money_cell(
            juniors.cell(row=row_index, column=6),
            item["team_leader_income"],
        )

    for letter, width in (
        ("A", 28),
        ("B", 12),
        ("C", 8),
        ("D", 14),
        ("E", 14),
        ("F", 14),
    ):
        juniors.column_dimensions[letter].width = width

    ops_sheet = workbook.create_sheet("Operations")
    op_headers = [
        "ID",
        "Date",
        "Agent",
        "Type",
        "Property",
        "Commission",
        "Agent payment",
    ]
    for index, header in enumerate(op_headers, start=1):
        _header_cell(ops_sheet.cell(row=1, column=index), header)

    for row_index, operation in enumerate(report["operations"], start=2):
        alt = row_index % 2 == 0
        _text_cell(ops_sheet.cell(row=row_index, column=1), operation.get("id"), alt)
        _text_cell(ops_sheet.cell(row=row_index, column=2), operation.get("date"), alt)
        _text_cell(ops_sheet.cell(row=row_index, column=3), operation.get("agent"), alt)
        _text_cell(ops_sheet.cell(row=row_index, column=4), operation.get("agent_type"), alt)
        _text_cell(ops_sheet.cell(row=row_index, column=5), operation.get("property"), alt)
        _money_cell(
            ops_sheet.cell(row=row_index, column=6),
            operation.get("total_commission"),
        )
        _money_cell(
            ops_sheet.cell(row=row_index, column=7),
            operation.get("agent_payment"),
        )

    for letter, width in (
        ("A", 14),
        ("B", 12),
        ("C", 22),
        ("D", 10),
        ("E", 28),
        ("F", 14),
        ("G", 14),
    ):
        ops_sheet.column_dimensions[letter].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

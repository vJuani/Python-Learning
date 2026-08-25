"""
RE/MAX Red export adapter (XLSX/CSV) → sync engine DTOs.

Does not create a second persistence pipeline: parsed rows become the
same ExternalProperty payload consumed by CsvUploadAdapter / sync.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Optional

from openpyxl import load_workbook

from modules.database.csv_import_batches_repository import (
    create_csv_import_batch,
    delete_csv_import_batch,
    get_csv_import_batch,
    update_csv_import_batch,
)
from modules.database.organization_integrations_repository import (
    PROVIDER_CSV_UPLOAD,
    SCOPE_AGENT,
    STATUS_CONNECTED,
    create_organization_integration,
    find_organization_integration_by_provider,
    update_organization_integration_config,
)
from modules.integrations.csv_import import (
    payload_to_external,
)
from modules.integrations.matching import (
    match_listing_by_external_id,
)
from modules.property_types import (
    LISTING_PURPOSES,
    PROPERTY_TYPES,
)
from modules.validators import JURISDICTIONS


DEFAULT_LISTING_PROVIDER = "remax_web"
MAX_EXPORT_ROWS = 5000
SUPPORTED_CURRENCIES = ("USD", "ARS")
PERCENT_MIN = Decimal("0")
PERCENT_MAX = Decimal("100")

REQUIRED_HEADER_ALIASES = {
    "mlsid": ("mlsid", "mls id", "mls_id"),
    "direccion": ("direccion", "dirección", "address", "calle"),
    "altura": ("altura",),
    "localidad": ("localidad", "ciudad", "city"),
    "status": (
        "status listing",
        "status_listing",
        "estado listing",
        "estado",
    ),
    "operacion": (
        "tipo de operacion",
        "tipo de operación",
        "operacion",
        "operación",
    ),
    "tipo_propiedad": (
        "tipo de propiedad",
        "tipo propiedad",
        "property type",
    ),
    "precio": ("precio", "price"),
    "moneda": (
        "tipo de moneda",
        "moneda",
        "currency",
    ),
}

OPTIONAL_HEADER_ALIASES = {
    "codigo_postal": (
        "codigo postal",
        "código postal",
        "codigo_postal",
        "cp",
        "postal_code",
    ),
    "pct_comprador": (
        "% al comprador",
        "% comprador",
        "pct comprador",
        "porcentaje comprador",
    ),
    "pct_vendedor": (
        "% al vendedor",
        "% vendedor",
        "pct vendedor",
        "porcentaje vendedor",
    ),
}

STATUS_ALIASES = {
    "activa": "active",
    "active": "active",
    "reservada": "reserved",
    "reserved": "reserved",
    "negociacion": "negotiation",
    "negociaciòn": "negotiation",
    "negotiation": "negotiation",
}

PURPOSE_ALIASES = {
    "venta": "sale",
    "sale": "sale",
    "alquiler": "rental",
    "rental": "rental",
    "alquiler temporario": "temporary_rental",
    "temporario": "temporary_rental",
    "temporary_rental": "temporary_rental",
}

PROPERTY_TYPE_ALIASES = {
    "departamento": "apartment",
    "departamento estandar": "apartment",
    "departamento estándar": "apartment",
    "departamento semipiso": "apartment",
    "departamento piso": "apartment",
    "departamento monoambiente": "apartment",
    "monoambiente": "apartment",
    "ph": "ph",
    "casa": "house",
    "casa duplex": "house",
    "casa dúplex": "house",
    "casa triplex": "house",
    "duplex": "house",
    "dúplex": "house",
    "triplex": "house",
    "terrenos y lotes": "land",
    "terreno": "land",
    "terrenos": "land",
    "lote": "land",
    "lotes": "land",
    "local": "commercial",
    "galpon": "commercial",
    "galpón": "commercial",
    "deposito": "commercial",
    "depósito": "commercial",
    "oficina": "office",
    "consultorio": "office",
}

CABA_LOCALITIES = frozenset(
    {
        "caba",
        "c.a.b.a.",
        "c.a.b.a",
        "capital federal",
        "ciudad autonoma de buenos aires",
        "ciudad autónoma de buenos aires",
        "ciudad de buenos aires",
        "buenos aires capital",
        "palermo",
        "belgrano",
        "recoleta",
        "caballito",
        "flores",
        "villa urquiza",
        "nunez",
        "nuñez",
        "núñez",
        "puerto madero",
        "san telmo",
        "retiro",
        "monserrat",
        "montserrat",
        "balvanera",
        "almagro",
        "villa crespo",
        "colegiales",
        "chacarita",
        "saavedra",
        "villa del parque",
        "villa devoto",
        "villa lugano",
        "mataderos",
        "liniers",
        "versalles",
        "villa luro",
        "floresta",
        "monte castro",
        "villa santa rita",
        "villa general mitre",
        "parque chas",
        "agronomia",
        "agronomía",
        "parque patricios",
        "barracas",
        "la boca",
        "constitucion",
        "constitución",
        "san nicolas",
        "san nicolás",
        "san cristobal",
        "san cristóbal",
        "boedo",
        "parque chacabuco",
        "nueva pompeya",
        "villa soldati",
        "villa riachuelo",
        "coghlan",
        "villa ortuzar",
        "villa ortúzar",
        "paternal",
        "villa pueyrredon",
        "villa pueyrredón",
        "coghlan",
    }
)

PBA_LOCALITIES = frozenset(
    {
        "la plata",
        "vicente lopez",
        "vicente lópez",
        "san isidro",
        "tigre",
        "pilar",
        "moron",
        "morón",
        "quilmes",
        "avellaneda",
        "lanus",
        "lanús",
        "lomas de zamora",
        "banfield",
        "martinez",
        "martínez",
        "olivos",
        "san fernando",
        "escobar",
        "san miguel",
        "ituzaingo",
        "ituzaingó",
        "hurlingham",
        "tres de febrero",
        "caseros",
        "san martin",
        "san martín",
        "villa ballester",
        "florida",
        "munro",
        "beccar",
        "beccar",
        "acassuso",
        "boulogne",
        "don torcuato",
        "general pacheco",
        "benavidez",
        "nordelta",
        "berazategui",
        "florencio varela",
        "almirante brown",
        "adrogue",
        "adrogué",
        "temperley",
        "banfield",
        "ramos mejia",
        "ramos mejía",
        "haedo",
        "castelar",
        "moreno",
        "merlo",
        "jose c. paz",
        "josé c. paz",
        "malvinas argentinas",
        "ezeiza",
        "esteban echeverria",
        "esteban echeverría",
        "san vicente",
        "berisso",
        "ensenada",
        "bernal",
        "wilde",
        "sarandi",
        "sarandí",
        "gerli",
        "remedios de escalada",
    }
)


@dataclass
class RemaxIssue:
    severity: str
    code: str
    message: str
    row_number: Optional[int] = None
    mlsid: Optional[str] = None


@dataclass
class RemaxSourceRow:
    row_number: int
    mlsid: str
    address: str
    locality: str
    postal_code: str
    status_raw: str
    status: Optional[str]
    property_type_raw: str
    property_type: Optional[str]
    purpose_raw: str
    listing_purpose: Optional[str]
    price: Optional[float]
    currency: Optional[str]
    buyer_side_commission_percent: Optional[float]
    seller_side_commission_percent: Optional[float]
    jurisdiction: Optional[str]
    jurisdiction_source: str
    url: Optional[str] = None
    blockers: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return len(self.blockers) == 0


@dataclass
class RemaxParseResult:
    rows: list[RemaxSourceRow] = field(default_factory=list)
    issues: list[RemaxIssue] = field(default_factory=list)
    filename: Optional[str] = None
    agent_id: Optional[int] = None
    agent_name: Optional[str] = None

    @property
    def has_blocking_errors(self) -> bool:
        if any(issue.severity == "error" for issue in self.issues):
            return True

        return any(not row.is_valid for row in self.rows)


def _issue(severity, code, message, row_number=None, mlsid=None):
    return RemaxIssue(
        severity=severity,
        code=code,
        message=message,
        row_number=row_number,
        mlsid=mlsid,
    )


def _strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(
        char
        for char in normalized
        if not unicodedata.combining(char)
    )


def normalize_text(value) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_key(value: str) -> str:
    text = _strip_accents(normalize_text(value)).lower()
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _map_headers(fieldnames) -> dict:
    normalized_to_original = {}

    for name in fieldnames:
        if name is None:
            continue
        key = normalize_key(str(name))
        if key:
            normalized_to_original[key] = str(name)

    mapping = {}

    for logical, aliases in {
        **REQUIRED_HEADER_ALIASES,
        **OPTIONAL_HEADER_ALIASES,
    }.items():
        for alias in aliases:
            alias_key = normalize_key(alias)
            if alias_key in normalized_to_original:
                mapping[logical] = normalized_to_original[
                    alias_key
                ]
                break

    return mapping


def _cell(row: dict, mapping: dict, logical: str) -> str:
    original = mapping.get(logical)
    if original is None:
        return ""
    return normalize_text(row.get(original))


def resolve_jurisdiction(
    locality: str,
    postal_code: str,
    *,
    override: Optional[str] = None,
) -> tuple[Optional[str], str]:
    if override in JURISDICTIONS:
        return override, "override"

    postal = normalize_text(postal_code).upper().replace(" ", "")
    locality_key = normalize_key(locality)

    from_cp = None
    if postal.startswith("C"):
        from_cp = "CABA"
    elif postal.startswith("B"):
        from_cp = "PBA"

    from_locality = None
    if locality_key in CABA_LOCALITIES:
        from_locality = "CABA"
    elif locality_key in PBA_LOCALITIES:
        from_locality = "PBA"

    if from_cp and from_locality and from_cp != from_locality:
        return None, "conflict"

    if from_cp:
        return from_cp, "postal_code"

    if from_locality:
        return from_locality, "locality"

    return None, "ambiguous"


def normalize_status(raw: str) -> Optional[str]:
    key = normalize_key(raw)
    return STATUS_ALIASES.get(key)


def normalize_purpose(raw: str) -> Optional[str]:
    key = normalize_key(raw)
    return PURPOSE_ALIASES.get(key)


def normalize_property_type(raw: str) -> Optional[str]:
    key = normalize_key(raw)

    if not key:
        return None

    if key in PROPERTY_TYPE_ALIASES:
        return PROPERTY_TYPE_ALIASES[key]

    if key.startswith("departamento"):
        return "apartment"

    if key.startswith("casa"):
        return "house"

    if "terreno" in key or "lote" in key:
        return "land"

    return None


def normalize_currency(raw: str) -> Optional[str]:
    key = normalize_key(raw).replace("$", "").strip()

    if key in ("usd", "u$s", "us$", "dolar", "dolares", "dólar", "dólares"):
        return "USD"

    if key in ("ars", "peso", "pesos", "$ar", "arg"):
        return "ARS"

    upper = normalize_text(raw).upper()
    if upper in SUPPORTED_CURRENCIES:
        return upper

    return None


def parse_percent(raw: str) -> tuple[Optional[Decimal], Optional[str]]:
    text = normalize_text(raw)

    if text == "":
        return None, None

    text = text.replace("%", "").strip()
    text = text.replace(" ", "")

    if "," in text and "." in text:
        # Ambiguous thousand/decimal mix — reject.
        return None, "percent_ambiguous_format"

    if "," in text:
        text = text.replace(",", ".")

    try:
        value = Decimal(text)
    except InvalidOperation:
        return None, "percent_invalid"

    if value < PERCENT_MIN or value > PERCENT_MAX:
        return None, "percent_out_of_range"

    return value, None


def parse_price(raw: str) -> tuple[Optional[Decimal], Optional[str]]:
    text = normalize_text(raw)

    if text == "":
        return None, None

    text = text.replace("$", "").replace(" ", "")

    if "," in text and "." in text:
        # 1.234.567,89 or 1,234,567.89
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")

    try:
        value = Decimal(text)
    except InvalidOperation:
        return None, "price_invalid"

    if value < 0:
        return None, "price_negative"

    return value, None


def build_address(street: str, height: str) -> str:
    street = normalize_text(street)
    height = normalize_text(height)

    if street and height:
        return f"{street} {height}"

    return street or height


def _decimal_to_float(value: Optional[Decimal]) -> Optional[float]:
    if value is None:
        return None
    return float(value)


def _rows_from_csv_text(text: str) -> tuple[list[str], list[dict]]:
    reader = csv.DictReader(io.StringIO(text))

    if reader.fieldnames is None:
        return [], []

    fieldnames = list(reader.fieldnames)
    rows = []

    for raw in reader:
        rows.append(
            {
                key: raw.get(key)
                for key in fieldnames
                if key is not None
            }
        )

    return fieldnames, rows


def _rows_from_xlsx(raw: bytes) -> tuple[list[str], list[dict]]:
    workbook = load_workbook(
        filename=io.BytesIO(raw),
        read_only=True,
        data_only=True,
    )
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        workbook.close()
        return [], []

    fieldnames = [
        "" if cell is None else str(cell)
        for cell in header_row
    ]

    rows = []
    for values in rows_iter:
        if values is None:
            continue
        item = {}
        empty = True
        for index, name in enumerate(fieldnames):
            value = values[index] if index < len(values) else None
            item[name] = value
            if value is not None and str(value).strip() != "":
                empty = False
        if not empty:
            rows.append(item)

    workbook.close()
    return fieldnames, rows


def _decode_csv_bytes(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("csv_encoding_invalid")


def load_tabular_rows(
    raw: bytes,
    *,
    filename: Optional[str] = None,
) -> tuple[list[str], list[dict], list[RemaxIssue]]:
    issues = []
    name = (filename or "").lower()

    if name.endswith(".xlsx"):
        try:
            fieldnames, rows = _rows_from_xlsx(raw)
        except Exception:
            issues.append(
                _issue(
                    "error",
                    "xlsx_unreadable",
                    "No se pudo leer el archivo XLSX.",
                )
            )
            return [], [], issues
        return fieldnames, rows, issues

    try:
        text = _decode_csv_bytes(raw)
    except ValueError:
        issues.append(
            _issue(
                "error",
                "csv_encoding_invalid",
                "No se pudo leer el CSV.",
            )
        )
        return [], [], issues

    fieldnames, rows = _rows_from_csv_text(text)
    return fieldnames, rows, issues


def _parse_rows(
    fieldnames: list[str],
    raw_rows: list[dict],
    *,
    filename: Optional[str] = None,
    overrides: Optional[dict] = None,
) -> RemaxParseResult:
    result = RemaxParseResult(filename=filename)
    overrides = overrides or {}

    if not fieldnames:
        result.issues.append(
            _issue(
                "error",
                "missing_header",
                "Falta la fila de encabezados.",
            )
        )
        return result

    mapping = _map_headers(fieldnames)

    for logical in REQUIRED_HEADER_ALIASES:
        if logical not in mapping:
            result.issues.append(
                _issue(
                    "error",
                    "missing_column",
                    f"Falta la columna obligatoria para '{logical}'.",
                )
            )

    if result.has_blocking_errors:
        return result

    seen_mlsids = {}
    data_count = 0

    for index, raw_row in enumerate(raw_rows, start=2):
        if not any(
            normalize_text(value) != ""
            for value in raw_row.values()
        ):
            continue

        data_count += 1

        if data_count > MAX_EXPORT_ROWS:
            result.issues.append(
                _issue(
                    "error",
                    "too_many_rows",
                    f"El archivo supera el límite de {MAX_EXPORT_ROWS} filas.",
                )
            )
            break

        mlsid = _cell(raw_row, mapping, "mlsid")
        street = _cell(raw_row, mapping, "direccion")
        height = _cell(raw_row, mapping, "altura")
        locality = _cell(raw_row, mapping, "localidad")
        postal_code = _cell(raw_row, mapping, "codigo_postal")
        status_raw = _cell(raw_row, mapping, "status")
        purpose_raw = _cell(raw_row, mapping, "operacion")
        type_raw = _cell(raw_row, mapping, "tipo_propiedad")
        price_raw = _cell(raw_row, mapping, "precio")
        currency_raw = _cell(raw_row, mapping, "moneda")
        buyer_raw = _cell(raw_row, mapping, "pct_comprador")
        seller_raw = _cell(raw_row, mapping, "pct_vendedor")

        row_overrides = overrides.get(mlsid) or {}
        blockers = []
        warnings = []

        if not mlsid:
            blockers.append("MLSID es obligatorio.")
        elif mlsid in seen_mlsids:
            blockers.append(
                f"MLSID duplicado (también en fila {seen_mlsids[mlsid]})."
            )
        else:
            seen_mlsids[mlsid] = index

        address = build_address(street, height)
        if not address:
            blockers.append("Dirección/Altura es obligatoria.")

        status = normalize_status(status_raw)
        if not status_raw:
            blockers.append("Status Listing es obligatorio.")
        elif status is None:
            blockers.append(
                f"Status Listing no reconocido: '{status_raw}'."
            )

        listing_purpose = normalize_purpose(purpose_raw)
        if not purpose_raw:
            blockers.append("Tipo de Operación es obligatorio.")
        elif listing_purpose is None:
            blockers.append(
                f"Tipo de Operación no reconocido: '{purpose_raw}'."
            )
        elif listing_purpose not in LISTING_PURPOSES:
            blockers.append("listing_purpose inválido.")

        property_type_override = row_overrides.get(
            "property_type"
        )
        if property_type_override:
            property_type = property_type_override
            if property_type not in PROPERTY_TYPES:
                blockers.append(
                    f"property_type override inválido: '{property_type}'."
                )
                property_type = None
        else:
            property_type = normalize_property_type(type_raw)
            if not type_raw:
                blockers.append("Tipo de Propiedad es obligatorio.")
            elif property_type is None:
                blockers.append(
                    f"Tipo de Propiedad desconocido: '{type_raw}'. "
                    "Resolvelo en el preview."
                )

        price_decimal, price_error = parse_price(price_raw)
        if price_error == "price_invalid":
            blockers.append("Precio inválido.")
        elif price_error == "price_negative":
            blockers.append("Precio no puede ser negativo.")
        elif price_decimal is None:
            warnings.append("Precio vacío; se importará sin listing_price.")

        currency = normalize_currency(currency_raw)
        if not currency_raw:
            blockers.append("Tipo de moneda es obligatorio.")
        elif currency is None:
            blockers.append(
                f"Moneda no soportada: '{currency_raw}' (USD/ARS)."
            )

        buyer_decimal, buyer_error = parse_percent(buyer_raw)
        if buyer_error == "percent_invalid":
            blockers.append("% al Comprador inválido.")
        elif buyer_error == "percent_ambiguous_format":
            blockers.append(
                "% al Comprador con formato ambiguo."
            )
        elif buyer_error == "percent_out_of_range":
            blockers.append(
                "% al Comprador fuera de rango (0–100). "
                "No se asume conversión automática."
            )

        seller_decimal, seller_error = parse_percent(seller_raw)
        if seller_error == "percent_invalid":
            blockers.append("% al Vendedor inválido.")
        elif seller_error == "percent_ambiguous_format":
            blockers.append(
                "% al Vendedor con formato ambiguo."
            )
        elif seller_error == "percent_out_of_range":
            blockers.append(
                "% al Vendedor fuera de rango (0–100). "
                "No se asume conversión automática."
            )

        jurisdiction_override = row_overrides.get(
            "jurisdiction"
        )
        jurisdiction, jurisdiction_source = resolve_jurisdiction(
            locality,
            postal_code,
            override=jurisdiction_override,
        )

        if jurisdiction is None:
            if jurisdiction_source == "conflict":
                blockers.append(
                    "Jurisdicción contradictoria entre CP y localidad."
                )
            else:
                blockers.append(
                    "No se pudo derivar jurisdicción (CABA/PBA). "
                    "Corregila en el preview."
                )
        elif jurisdiction not in JURISDICTIONS:
            blockers.append("Jurisdicción inválida.")

        for message in blockers:
            result.issues.append(
                _issue(
                    "error",
                    "row_blocker",
                    message,
                    row_number=index,
                    mlsid=mlsid or None,
                )
            )

        for message in warnings:
            result.issues.append(
                _issue(
                    "warning",
                    "row_warning",
                    message,
                    row_number=index,
                    mlsid=mlsid or None,
                )
            )

        result.rows.append(
            RemaxSourceRow(
                row_number=index,
                mlsid=mlsid,
                address=address,
                locality=locality,
                postal_code=postal_code,
                status_raw=status_raw,
                status=status,
                property_type_raw=type_raw,
                property_type=property_type,
                purpose_raw=purpose_raw,
                listing_purpose=listing_purpose,
                price=_decimal_to_float(price_decimal),
                currency=currency,
                buyer_side_commission_percent=_decimal_to_float(
                    buyer_decimal
                ),
                seller_side_commission_percent=_decimal_to_float(
                    seller_decimal
                ),
                jurisdiction=jurisdiction,
                jurisdiction_source=jurisdiction_source,
                blockers=blockers,
                warnings=warnings,
            )
        )

    if data_count == 0 and not result.has_blocking_errors:
        result.issues.append(
            _issue(
                "error",
                "no_data_rows",
                "El archivo no tiene filas de datos.",
            )
        )

    return result


def parse_remax_export_bytes(
    raw: bytes,
    *,
    filename: Optional[str] = None,
    overrides: Optional[dict] = None,
) -> RemaxParseResult:
    fieldnames, raw_rows, load_issues = load_tabular_rows(
        raw,
        filename=filename,
    )

    if load_issues:
        result = RemaxParseResult(filename=filename)
        result.issues.extend(load_issues)
        return result

    return _parse_rows(
        fieldnames,
        raw_rows,
        filename=filename,
        overrides=overrides,
    )


def source_rows_to_payload(
    rows: list[RemaxSourceRow],
    *,
    agent_id: int,
    agent_name: str,
    agent_external_id: str,
    overrides: Optional[dict] = None,
    raw_file_b64: Optional[str] = None,
) -> dict:
    properties = []

    for row in rows:
        if not row.is_valid:
            continue

        properties.append(
            {
                "external_id": row.mlsid,
                "agent_external_id": agent_external_id,
                "address": row.address,
                "jurisdiction": row.jurisdiction,
                "url": row.url,
                "listing_provider": DEFAULT_LISTING_PROVIDER,
                "listing_status": row.status,
                "property_type": row.property_type,
                "listing_price": row.price,
                "listing_purpose": row.listing_purpose,
                "listing_currency": row.currency,
                "buyer_side_commission_percent": (
                    row.buyer_side_commission_percent
                ),
                "seller_side_commission_percent": (
                    row.seller_side_commission_percent
                ),
            }
        )

    return {
        "agents": [],
        "properties": properties,
        "meta": {
            "format": "remax_export",
            "agent_id": agent_id,
            "agent_name": agent_name,
            "agent_external_id": agent_external_id,
            "overrides": overrides or {},
            "source_rows": [
                {
                    "row_number": row.row_number,
                    "mlsid": row.mlsid,
                    "address": row.address,
                    "locality": row.locality,
                    "postal_code": row.postal_code,
                    "status_raw": row.status_raw,
                    "status": row.status,
                    "property_type_raw": row.property_type_raw,
                    "property_type": row.property_type,
                    "purpose_raw": row.purpose_raw,
                    "listing_purpose": row.listing_purpose,
                    "price": row.price,
                    "currency": row.currency,
                    "buyer_side_commission_percent": (
                        row.buyer_side_commission_percent
                    ),
                    "seller_side_commission_percent": (
                        row.seller_side_commission_percent
                    ),
                    "jurisdiction": row.jurisdiction,
                    "jurisdiction_source": row.jurisdiction_source,
                    "blockers": row.blockers,
                    "warnings": row.warnings,
                    "is_valid": row.is_valid,
                }
                for row in rows
            ],
            "raw_file_b64": raw_file_b64,
            "filename": None,
        },
    }


def build_remax_preview(
    organization_id,
    parse_result: RemaxParseResult,
    *,
    agent_id: int,
    agent_name: str,
    agent_external_id: str,
    overrides: Optional[dict] = None,
    raw_file_b64: Optional[str] = None,
) -> dict:
    payload = source_rows_to_payload(
        parse_result.rows,
        agent_id=agent_id,
        agent_name=agent_name,
        agent_external_id=agent_external_id,
        overrides=overrides,
        raw_file_b64=raw_file_b64,
    )
    payload["meta"]["filename"] = parse_result.filename

    issues = [asdict(issue) for issue in parse_result.issues]
    properties_preview = []
    properties_new = 0
    properties_update = 0

    for prop in payload["properties"]:
        matched = match_listing_by_external_id(
            organization_id,
            prop["listing_provider"],
            prop["external_id"],
        )
        action = "create" if matched is None else "update"

        if action == "create":
            properties_new += 1
        else:
            properties_update += 1

        source = next(
            (
                row
                for row in payload["meta"]["source_rows"]
                if row["mlsid"] == prop["external_id"]
            ),
            {},
        )

        properties_preview.append(
            {
                "property_external_id": prop["external_id"],
                "mlsid": prop["external_id"],
                "address": prop["address"],
                "locality": source.get("locality"),
                "jurisdiction": prop["jurisdiction"],
                "jurisdiction_source": source.get(
                    "jurisdiction_source"
                ),
                "status_raw": source.get("status_raw"),
                "status": prop["listing_status"],
                "property_type_raw": source.get(
                    "property_type_raw"
                ),
                "property_type": prop.get("property_type"),
                "listing_purpose": prop.get("listing_purpose"),
                "purpose_raw": source.get("purpose_raw"),
                "price": prop.get("listing_price"),
                "currency": prop.get("listing_currency"),
                "buyer_side_commission_percent": prop.get(
                    "buyer_side_commission_percent"
                ),
                "seller_side_commission_percent": prop.get(
                    "seller_side_commission_percent"
                ),
                "listing_provider": prop["listing_provider"],
                "url": prop.get("url"),
                "match": (
                    None
                    if matched is None
                    else {
                        "listing_id": matched["id"],
                        "property_id": matched["property_id"],
                    }
                ),
                "action": action,
            }
        )

    # Include invalid rows in preview table for resolution UI.
    valid_mlsids = {
        item["mlsid"] for item in properties_preview
    }
    for source in payload["meta"]["source_rows"]:
        if source["mlsid"] in valid_mlsids:
            continue
        properties_preview.append(
            {
                "property_external_id": source["mlsid"],
                "mlsid": source["mlsid"],
                "address": source["address"],
                "locality": source["locality"],
                "jurisdiction": source["jurisdiction"],
                "jurisdiction_source": source[
                    "jurisdiction_source"
                ],
                "status_raw": source["status_raw"],
                "status": source["status"],
                "property_type_raw": source[
                    "property_type_raw"
                ],
                "property_type": source["property_type"],
                "listing_purpose": source["listing_purpose"],
                "purpose_raw": source["purpose_raw"],
                "price": source["price"],
                "currency": source["currency"],
                "buyer_side_commission_percent": source[
                    "buyer_side_commission_percent"
                ],
                "seller_side_commission_percent": source[
                    "seller_side_commission_percent"
                ],
                "listing_provider": DEFAULT_LISTING_PROVIDER,
                "url": None,
                "match": None,
                "action": "blocked",
                "blockers": source["blockers"],
            }
        )

    blocking_errors = [
        issue
        for issue in issues
        if issue["severity"] == "error"
    ]
    warnings = [
        issue
        for issue in issues
        if issue["severity"] == "warning"
    ]

    can_confirm = (
        len(blocking_errors) == 0
        and len(payload["properties"]) > 0
        and all(
            row["is_valid"]
            for row in payload["meta"]["source_rows"]
        )
    )

    return {
        "filename": parse_result.filename,
        "format": "remax_export",
        "can_confirm": can_confirm,
        "default_listing_provider": DEFAULT_LISTING_PROVIDER,
        "listing_providers": [DEFAULT_LISTING_PROVIDER],
        "integration_provider": PROVIDER_CSV_UPLOAD,
        "agent": {
            "id": agent_id,
            "name": agent_name,
            "external_id": agent_external_id,
        },
        "property_types": list(PROPERTY_TYPES),
        "jurisdictions": list(JURISDICTIONS),
        "summary": {
            "rows_total": len(parse_result.rows),
            "rows_valid": len(payload["properties"]),
            "properties_new": properties_new,
            "properties_update": properties_update,
            "listings_new": properties_new,
            "listings_update": properties_update,
            "errors": len(blocking_errors),
            "warnings": len(warnings),
        },
        "properties": properties_preview,
        "issues": issues,
        "payload": payload,
    }


def stage_remax_import(
    organization_id,
    parse_result: RemaxParseResult,
    *,
    agent_id: int,
    agent_name: str,
    agent_external_id: str,
    overrides: Optional[dict] = None,
    raw_file_b64: Optional[str] = None,
):
    preview = build_remax_preview(
        organization_id,
        parse_result,
        agent_id=agent_id,
        agent_name=agent_name,
        agent_external_id=agent_external_id,
        overrides=overrides,
        raw_file_b64=raw_file_b64,
    )
    payload = preview.pop("payload")

    return create_csv_import_batch(
        organization_id,
        filename=parse_result.filename,
        payload=payload,
        preview=preview,
    )


def agent_external_id_for_local(agent: dict) -> str:
    external_id = (agent.get("external_id") or "").strip()
    if external_id:
        return external_id
    return f"local:{agent['id']}"


def prepare_remax_integration_for_batch(
    organization_id,
    batch_id,
):
    batch = get_csv_import_batch(batch_id, organization_id)

    if batch is None:
        raise ValueError("csv_batch_not_found")

    preview = batch["preview"]

    if not preview.get("can_confirm"):
        raise ValueError("csv_batch_has_blockers")

    meta = (batch["payload"] or {}).get("meta") or {}
    agent_id = meta.get("agent_id")

    if agent_id is None:
        raise ValueError("remax_agent_required")

    existing = find_organization_integration_by_provider(
        organization_id,
        PROVIDER_CSV_UPLOAD,
        scope_type=SCOPE_AGENT,
        agent_id=agent_id,
    )

    if existing is None:
        existing = create_organization_integration(
            organization_id,
            PROVIDER_CSV_UPLOAD,
            SCOPE_AGENT,
            agent_id=agent_id,
            status=STATUS_CONNECTED,
            external_office_id=f"remax-export-agent-{agent_id}",
            config={},
        )

    config = dict(existing.get("config") or {})
    config.update(
        {
            "batch_id": batch_id,
            "deactivate_missing_listings": False,
            "default_listing_provider": (
                DEFAULT_LISTING_PROVIDER
            ),
            "import_format": "remax_export",
            "original_filename": batch.get("filename"),
            "payload": {
                "agents": batch["payload"].get("agents", []),
                "properties": batch["payload"].get(
                    "properties",
                    [],
                ),
            },
        }
    )

    return update_organization_integration_config(
        existing["id"],
        organization_id,
        config,
    )


def clear_remax_batch_from_integration(
    organization_id,
    integration_id,
    batch_id,
):
    from modules.database.organization_integrations_repository import (
        get_organization_integration,
    )

    integration = get_organization_integration(
        integration_id,
        organization_id,
    )

    if integration is None:
        delete_csv_import_batch(batch_id, organization_id)
        return None

    config = dict(integration.get("config") or {})
    config.pop("batch_id", None)
    config.pop("payload", None)
    config["deactivate_missing_listings"] = False
    config["last_import_filename"] = config.pop(
        "original_filename",
        None,
    )
    config["import_format"] = "remax_export"

    updated = update_organization_integration_config(
        integration_id,
        organization_id,
        config,
    )
    delete_csv_import_batch(batch_id, organization_id)
    return updated


def _source_dict_to_row(item: dict) -> RemaxSourceRow:
    return RemaxSourceRow(
        row_number=item["row_number"],
        mlsid=item.get("mlsid") or "",
        address=item.get("address") or "",
        locality=item.get("locality") or "",
        postal_code=item.get("postal_code") or "",
        status_raw=item.get("status_raw") or "",
        status=item.get("status"),
        property_type_raw=item.get("property_type_raw") or "",
        property_type=item.get("property_type"),
        purpose_raw=item.get("purpose_raw") or "",
        listing_purpose=item.get("listing_purpose"),
        price=item.get("price"),
        currency=item.get("currency"),
        buyer_side_commission_percent=item.get(
            "buyer_side_commission_percent"
        ),
        seller_side_commission_percent=item.get(
            "seller_side_commission_percent"
        ),
        jurisdiction=item.get("jurisdiction"),
        jurisdiction_source=item.get(
            "jurisdiction_source"
        )
        or "ambiguous",
        blockers=list(item.get("blockers") or []),
        warnings=list(item.get("warnings") or []),
    )


def apply_remax_preview_overrides(
    organization_id,
    batch_id,
    overrides: dict,
):
    """
    Apply Admin jurisdiction / property_type overrides on a staged batch
    without re-reading the original file.
    """
    batch = get_csv_import_batch(batch_id, organization_id)

    if batch is None:
        raise ValueError("csv_batch_not_found")

    meta = (batch["payload"] or {}).get("meta") or {}
    agent_id = meta.get("agent_id")
    agent_name = meta.get("agent_name")
    agent_external_id = meta.get("agent_external_id")
    source_rows = meta.get("source_rows") or []

    if agent_id is None or not agent_name or not agent_external_id:
        raise ValueError("remax_agent_required")

    rebuilt_rows = []
    issues = []

    for item in source_rows:
        row = _source_dict_to_row(item)
        row_overrides = overrides.get(row.mlsid) or {}
        blockers = []
        warnings = list(row.warnings)

        # Keep previously non-override blockers except jurisdiction/type.
        for message in row.blockers:
            if (
                "jurisdicción" in message.lower()
                or "jurisdiction" in message.lower()
                or "tipo de propiedad" in message.lower()
                or "property_type" in message.lower()
            ):
                continue
            blockers.append(message)

        property_type_override = row_overrides.get(
            "property_type"
        )
        if property_type_override:
            if property_type_override in PROPERTY_TYPES:
                row.property_type = property_type_override
            else:
                blockers.append(
                    f"property_type override inválido: "
                    f"'{property_type_override}'."
                )
                row.property_type = None
        elif row.property_type is None:
            if not row.property_type_raw:
                blockers.append(
                    "Tipo de Propiedad es obligatorio."
                )
            else:
                blockers.append(
                    f"Tipo de Propiedad desconocido: "
                    f"'{row.property_type_raw}'. "
                    "Resolvelo en el preview."
                )

        jurisdiction_override = row_overrides.get(
            "jurisdiction"
        )
        jurisdiction, jurisdiction_source = resolve_jurisdiction(
            row.locality,
            row.postal_code,
            override=jurisdiction_override,
        )
        row.jurisdiction = jurisdiction
        row.jurisdiction_source = jurisdiction_source

        if jurisdiction is None:
            if jurisdiction_source == "conflict":
                blockers.append(
                    "Jurisdicción contradictoria entre CP y localidad."
                )
            else:
                blockers.append(
                    "No se pudo derivar jurisdicción (CABA/PBA). "
                    "Corregila en el preview."
                )

        if not row.mlsid:
            blockers.append("MLSID es obligatorio.")
        if not row.address:
            blockers.append("Dirección/Altura es obligatoria.")
        if row.status is None:
            blockers.append(
                f"Status Listing no reconocido: '{row.status_raw}'."
            )
        if row.listing_purpose is None:
            blockers.append(
                f"Tipo de Operación no reconocido: '{row.purpose_raw}'."
            )
        if row.currency is None:
            blockers.append("Tipo de moneda es obligatorio.")

        row.blockers = blockers

        for message in blockers:
            issues.append(
                _issue(
                    "error",
                    "row_blocker",
                    message,
                    row_number=row.row_number,
                    mlsid=row.mlsid or None,
                )
            )

        for message in warnings:
            issues.append(
                _issue(
                    "warning",
                    "row_warning",
                    message,
                    row_number=row.row_number,
                    mlsid=row.mlsid or None,
                )
            )

        rebuilt_rows.append(row)

    parse_result = RemaxParseResult(
        rows=rebuilt_rows,
        issues=issues,
        filename=batch.get("filename") or meta.get("filename"),
        agent_id=agent_id,
        agent_name=agent_name,
    )

    preview = build_remax_preview(
        organization_id,
        parse_result,
        agent_id=agent_id,
        agent_name=agent_name,
        agent_external_id=agent_external_id,
        overrides=overrides,
    )
    payload = preview.pop("payload")

    return update_csv_import_batch(
        batch_id,
        organization_id,
        payload=payload,
        preview=preview,
    )


# Re-export for adapter consumers that import payload helpers.
__all__ = [
    "agent_external_id_for_local",
    "apply_remax_preview_overrides",
    "build_remax_preview",
    "clear_remax_batch_from_integration",
    "normalize_currency",
    "normalize_property_type",
    "normalize_purpose",
    "normalize_status",
    "parse_percent",
    "parse_price",
    "parse_remax_export_bytes",
    "payload_to_external",
    "prepare_remax_integration_for_batch",
    "resolve_jurisdiction",
    "stage_remax_import",
]

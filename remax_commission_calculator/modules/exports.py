import csv
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from modules.branding import get_brand_name
from modules.dashboard import (
    calculate_dashboard_metrics,
    get_agent_ranking
)

from modules.data_manager import (
    load_agents,
    load_history,
    load_properties
)


def export_history_csv():
    history = load_history()

    if len(history) == 0:
        print("No operations available to export.")
        return

    file_name = "commission_history.csv"

    fields = [
        "id",
        "date",
        "agent",
        "agent_type",
        "property_id",
        "property",
        "jurisdiction",
        "was_invoiced",
        "vat_amount",
        "sale_price",
        "commission_rate",
        "total_commission",
        "commission_after_abao",
        "abao",
        "martillero",
        "agent_payment",
        "office_payment",
        "office_total"
    ]

    with open(
        file_name,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fields
        )

        writer.writeheader()

        for operation in history:
            writer.writerow(operation)

    print(
        f"\nHistory exported successfully to "
        f"{file_name}"
    )


def export_excel(
    operations,
    file_name
):
    if len(operations) == 0:
        print(
            "No operations available to export."
        )
        return

    agents = load_agents()
    properties = load_properties()

    metrics = calculate_dashboard_metrics(
        operations
    )

    sorted_agents = get_agent_ranking(
        operations
    )

    workbook = Workbook()

    # =========================================
    # OPERATIONS SHEET
    # =========================================

    worksheet = workbook.active
    worksheet.title = "Operations"

    fields = [
        "ID",
        "Date",
        "Agent",
        "Agent Type",
        "Property ID",
        "Property",
        "Jurisdiction",
        "Invoiced",
        "VAT",
        "Sale Price",
        "Commission Rate",
        "Total Commission",
        "Commission After ABAO",
        "ABAO",
        "Martillero",
        "Agent Payment",
        "Office Payment",
        "Office Total"
    ]

    worksheet.append(fields)

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for operation in operations:
        worksheet.append([
            operation["id"],
            operation["date"],
            operation["agent"],
            operation["agent_type"],
            operation["property_id"],
            operation["property"],
            operation["jurisdiction"],
            operation["was_invoiced"],
            operation["vat_amount"],
            operation["sale_price"],
            operation["commission_rate"],
            operation["total_commission"],
            operation["commission_after_abao"],
            operation["abao"],
            operation["martillero"],
            operation["agent_payment"],
            operation["office_payment"],
            operation["office_total"]
        ])

    # =========================================
    # COLUMN WIDTHS
    # =========================================

    for column in range(
        1,
        worksheet.max_column + 1
    ):
        max_length = 0

        column_letter = get_column_letter(
            column
        )

        for cell in worksheet[column_letter]:
            if cell.value is not None:
                value_length = len(
                    str(cell.value)
                )

                if value_length > max_length:
                    max_length = value_length

        worksheet.column_dimensions[
            column_letter
        ].width = max_length + 4

    # =========================================
    # MONEY FORMAT
    # =========================================

    money_columns = [
        "I",
        "J",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R"
    ]

    for column in money_columns:
        for cell in worksheet[column][1:]:
            cell.number_format = (
                '"USD " #,##0.00'
            )

    for cell in worksheet["K"][1:]:
        cell.number_format = (
            '0.00"%"'
        )

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    # =========================================
    # DASHBOARD SHEET
    # =========================================

    dashboard = workbook.create_sheet(
        title="Dashboard",
        index=0
    )

    dashboard["A1"] = f"{get_brand_name()} Report"

    dashboard["A1"].font = Font(
        bold=True,
        size=18
    )

    dashboard["A3"] = "Operations"
    dashboard["B3"] = metrics[
        "total_operations"
    ]

    dashboard["A4"] = "Agents"
    dashboard["B4"] = len(agents)

    dashboard["A5"] = "Properties"
    dashboard["B5"] = len(properties)

    dashboard["A7"] = (
        "Gross Commission"
    )

    dashboard["B7"] = metrics[
        "gross_commission"
    ]

    dashboard["A8"] = (
        "Office Revenue"
    )

    dashboard["B8"] = metrics[
        "office_revenue"
    ]

    dashboard["A9"] = (
        "Agent Payments"
    )

    dashboard["B9"] = metrics[
        "agent_payments"
    ]

    dashboard["A10"] = (
        "Highest Commission"
    )

    dashboard["B10"] = metrics[
        "highest_commission"
    ]

    dashboard["A11"] = (
        "Average Commission"
    )

    dashboard["B11"] = metrics[
        "average_commission"
    ]

    for row in range(3, 12):
        dashboard[
            f"A{row}"
        ].font = Font(
            bold=True
        )

    for cell in [
        "B7",
        "B8",
        "B9",
        "B10",
        "B11"
    ]:
        dashboard[
            cell
        ].number_format = (
            '"USD " #,##0.00'
        )

    # =========================================
    # TOP AGENTS
    # =========================================

    dashboard["A14"] = (
        "TOP AGENTS"
    )

    dashboard["A14"].font = Font(
        bold=True,
        size=14
    )

    dashboard["A15"] = "Agent"
    dashboard["B15"] = "Commission"

    dashboard["A15"].font = Font(
        bold=True
    )

    dashboard["B15"].font = Font(
        bold=True
    )

    for index, (
        agent_name,
        commission
    ) in enumerate(
        sorted_agents[:3],
        start=16
    ):
        dashboard[
            f"A{index}"
        ] = agent_name

        dashboard[
            f"B{index}"
        ] = commission

        dashboard[
            f"B{index}"
        ].number_format = (
            '"USD " #,##0.00'
        )

    dashboard.column_dimensions[
        "A"
    ].width = 24

    dashboard.column_dimensions[
        "B"
    ].width = 20

    # =========================================
    # TOP AGENTS CHART
    # =========================================

    if len(sorted_agents) > 0:
        top_agents_count = min(
            3,
            len(sorted_agents)
        )

        chart = BarChart()

        chart.title = (
            "Top Agents by Commission"
        )

        chart.y_axis.title = (
            "Agent"
        )

        chart.x_axis.title = (
            "Commission (USD)"
        )

        data = Reference(
            dashboard,
            min_col=2,
            min_row=15,
            max_row=(
                15 + top_agents_count
            )
        )

        categories = Reference(
            dashboard,
            min_col=1,
            min_row=16,
            max_row=(
                15 + top_agents_count
            )
        )

        chart.add_data(
            data,
            titles_from_data=True
        )

        chart.set_categories(
            categories
        )

        chart.height = 7
        chart.width = 12

        dashboard.add_chart(
            chart,
            "D3"
        )

    # =========================================
    # COMMISSION BY OPERATION CHART
    # =========================================

    if len(operations) > 0:
        operations_chart = BarChart()

        operations_chart.title = (
            "Commission by Operation"
        )

        operations_chart.y_axis.title = (
            "Commission (USD)"
        )

        operations_chart.x_axis.title = (
            "Operation"
        )

        data = Reference(
            worksheet,
            min_col=12,
            min_row=1,
            max_row=worksheet.max_row
        )

        categories = Reference(
            worksheet,
            min_col=1,
            min_row=2,
            max_row=worksheet.max_row
        )

        operations_chart.add_data(
            data,
            titles_from_data=True
        )

        operations_chart.set_categories(
            categories
        )

        operations_chart.height = 8
        operations_chart.width = 16

        dashboard.add_chart(
            operations_chart,
            "D18"
        )

    # =========================================
    # SAVE FILE
    # =========================================

    workbook.save(
        file_name
    )

    print(
        f"\nExcel exported successfully to "
        f"{file_name}"
    )


def export_history_excel():
    history = load_history()

    export_excel(
        history,
        "commission_history.xlsx"
    )


def export_agent_report():
    history = load_history()

    if len(history) == 0:
        print(
            "No operations available to export."
        )
        return

    agent_search = input(
        "\nEnter the agent's name: "
    ).strip()

    if agent_search == "":
        print("Search cannot be empty.")
        return

    agent_operations = [
        operation
        for operation in history
        if agent_search.lower()
        in operation["agent"].lower()
    ]

    if len(agent_operations) == 0:
        print(
            f"No operations found for agent: "
            f"{agent_search}"
        )
        return

    file_name = (
        agent_search
        .lower()
        .replace(" ", "_")
        + "_report.xlsx"
    )

    export_excel(
        agent_operations,
        file_name
    )


def export_property_report():
    history = load_history()

    if len(history) == 0:
        print(
            "No operations available to export."
        )
        return

    property_search = input(
        "\nEnter the property address: "
    ).strip()

    if property_search == "":
        print("Search cannot be empty.")
        return

    property_operations = [
        operation
        for operation in history
        if property_search.lower()
        in operation["property"].lower()
    ]

    if len(property_operations) == 0:
        print(
            f"No operations found for property: "
            f"{property_search}"
        )
        return

    file_name = (
        property_search
        .lower()
        .replace(" ", "_")
        + "_report.xlsx"
    )

    export_excel(
        property_operations,
        file_name
    )


def export_date_report():
    history = load_history()

    if len(history) == 0:
        print(
            "No operations available to export."
        )
        return

    date_search = input(
        "\nEnter the date (dd/mm/yyyy): "
    ).strip()

    try:
        datetime.strptime(
            date_search,
            "%d/%m/%Y"
        )

    except ValueError:
        print(
            "Invalid date format. "
            "Use dd/mm/yyyy."
        )
        return

    date_operations = [
        operation
        for operation in history
        if operation["date"] == date_search
    ]

    if len(date_operations) == 0:
        print(
            f"No operations found for date: "
            f"{date_search}"
        )
        return

    safe_date = date_search.replace(
        "/",
        "-"
    )

    file_name = (
        f"{safe_date}_report.xlsx"
    )

    export_excel(
        date_operations,
        file_name
    )